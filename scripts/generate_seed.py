"""
Generate seed_data.sql from real Excel data + realistic demo values
for fields not present in the Excel file.

Excluded fields (user request): marital_status, num_children, youngest_child_age

Generated demo fields (deterministic via seeded random):
  gender, date_of_birth, country_id, email,
  entry_gpa, funding_id, has_external_work, weekly_work_hours,
  in_research_group, family_support, is_cross_discipline

Usage:  python scripts/generate_seed.py
Output: scripts/seed_data.sql
"""
import sys, os, re, random
from datetime import date, datetime, timedelta
from collections import defaultdict

sys.stdout.reconfigure(encoding="utf-8")

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl not installed. Run: python -m pip install openpyxl")
    sys.exit(1)

EXCEL_PATH  = r"D:\CodingSpace\DATATRAIN\PHD_data - Copy.xlsx"
OUTPUT_PATH = os.path.join(os.path.dirname(__file__), "seed_data.sql")
TODAY       = date(2026, 4, 28)
RNG         = random.Random(42)   # fixed seed → reproducible

# ══════════════════════════════════════════════════════════════════════
# LOOKUP MAPS
# ══════════════════════════════════════════════════════════════════════
CAMPUS_MAP = {"CYBER": 1, "MLAKA": 2}

PROGRAM_MAP = {
    "Ph.D. (Eng)":           (1, "PhD"),
    "Ph.D. (Mgmt.)":         (2, "PhD"),
    "Master (Mgmt)":         (3, "Master"),
    "Ph.D. (Communication)": (4, "PhD"),
    "Master (Communication)":(5, "Master"),
    "Master (Law)":          (6, "Master"),
}

STUDY_METHOD_MAP = {"full-time": "Full-time", "part-time": "Part-time"}

# Country IDs (match INSERT order below)
# 1=China 2=Japan 3=SouthKorea 4=India 5=Pakistan 6=Bangladesh
# 7=Malaysia 8=Indonesia 9=Thailand 10=Vietnam
# 11=SaudiArabia 12=Iran 13=UK 14=Germany 15=Nigeria 16=Kenya
# 17=USA 18=Brazil
COUNTRY_MALAYSIA = 7
COUNTRY_CHINA    = 1
COUNTRY_INDIA    = 4
COUNTRY_PAKISTAN = 5
COUNTRY_BANGLA   = 6
COUNTRY_NIGERIA  = 15

# Funding IDs: 1=Full Scholarship 2=Partial 3=Self-funded 4=TA 5=RA 6=Employer
FUND_FULL   = 1
FUND_PART   = 2
FUND_SELF   = 3
FUND_TA     = 4
FUND_RA     = 5
FUND_EMPLOY = 6

# ══════════════════════════════════════════════════════════════════════
# HELPERS
# ══════════════════════════════════════════════════════════════════════
def d(val):
    if val is None: return "NULL"
    if isinstance(val, datetime): val = val.date()
    return f"'{val.strftime('%Y-%m-%d')}'"

def q(val):
    if val is None: return "NULL"
    return f"'{str(val).replace(chr(39), chr(39)+chr(39))}'"

def b(val):
    return "TRUE" if val else "FALSE"

def milestone_status(expected, actual):
    if actual is not None: return "Completed"
    if expected is None:   return "Pending"
    exp = expected.date() if isinstance(expected, datetime) else expected
    return "Overdue" if exp <= TODAY else "Pending"

def milestone_actual_status(expected, actual):
    """Only mark Completed if actual date is in the past (real completion)."""
    if actual is None:
        exp = expected.date() if isinstance(expected, datetime) else expected
        if exp is None: return "Pending"
        return "Overdue" if exp <= TODAY else "Pending"
    act = actual.date() if isinstance(actual, datetime) else actual
    if act <= TODAY:
        return "Completed"
    # actual date is in the future → still pending
    return "Pending"


# ══════════════════════════════════════════════════════════════════════
# NAME-BASED INFERENCE HELPERS
# ══════════════════════════════════════════════════════════════════════
_MALAY_MARKERS  = {"BINTI", "BIN", "BTE", "BINT"}
_CHINESE_SURNAMES = {"WONG", "TAN", "LIM", "GOH", "YAP", "ONG", "KOH",
                     "ANG", "LEE", "CHAN", "CHONG", "NG", "YEOH", "LEONG",
                     "HANG", "YONG"}
_CHINESE_NAMES   = {"YANG", "XIAO"}
_INDIAN_MARKERS  = {"A/L", "A/P"}
_INDIAN_NAMES    = {"KUMAR", "SINGH", "NAIR", "DEVI", "PRASAD",
                    "JAYAKANTHAN", "JOANNA", "PUARA", "RAJESH", "PRIYA"}
_SOUTH_ASIAN     = {"KHAN", "MUHAMMAD", "HAMZA", "NAHIDA", "ANUM",
                    "NAZ", "MD", "UDDIN", "FERDOUS", "PIPUL", "BARKA"}
_NIGERIAN        = {"USMAN", "YAU", "IDRIS", "NUHU", "DAN-AZUMI", "OLALEKAN"}

def infer_country(name: str) -> int:
    words = set(re.split(r"[\s/]+", name.upper()))
    if words & _MALAY_MARKERS:       return COUNTRY_MALAYSIA
    if words & _INDIAN_MARKERS:      return COUNTRY_MALAYSIA   # Malaysian Indian
    if words & _CHINESE_SURNAMES:    return COUNTRY_MALAYSIA   # Malaysian Chinese
    if words & _CHINESE_NAMES:       return COUNTRY_CHINA
    if words & _INDIAN_NAMES:        return COUNTRY_INDIA
    if words & _SOUTH_ASIAN:
        return RNG.choice([COUNTRY_PAKISTAN, COUNTRY_BANGLA])
    if words & _NIGERIAN:            return COUNTRY_NIGERIA
    return RNG.choices(
        [COUNTRY_MALAYSIA, COUNTRY_CHINA, COUNTRY_INDIA, COUNTRY_PAKISTAN],
        weights=[40, 25, 20, 15],
    )[0]

_FEMALE_NAMES = {"NAHIDA", "ANUM", "NAZ", "NUR", "SITI", "AISHA", "JOANNA",
                 "PUARA", "PRIYA", "NURUL", "FATIMAH", "MIRA", "YAP", "LIM",
                 "HANG"}  # Hang Jia Yee = female

def infer_gender(name: str) -> str:
    upper = name.upper()
    # "A/L" = anak lelaki (son of) → Male; "A/P" = anak perempuan → Female
    if " A/L " in upper: return "Male"
    if " A/P " in upper: return "Female"
    first = name.strip().split()[0].upper()
    if first in _FEMALE_NAMES: return "Female"
    # common male prefixes / names
    if first in {"MUHAMMAD", "HAMZA", "AHMAD", "MD", "YANG", "USMAN",
                 "IDRIS", "NUHU", "KHAN", "RAHUL", "RAJESH", "ONG",
                 "WONG", "TAN", "GOH", "KOH", "ANG", "CHAN", "CHONG",
                 "JAYAKANTHAN", "FARHAN", "MOHAMMAD", "ABDUL", "CHONG"}:
        return "Male"
    return RNG.choice(["Male", "Male", "Female"])   # 2:1 male bias

def gen_dob(enrollment_date, degree_type: str) -> date:
    """Generate a plausible date of birth based on degree and enrollment date."""
    if isinstance(enrollment_date, datetime):
        enrollment_date = enrollment_date.date()
    if degree_type == "PhD":
        age = RNG.randint(24, 34)
    else:
        age = RNG.randint(22, 30)
    approx = enrollment_date - timedelta(days=age * 365 + RNG.randint(-180, 180))
    return approx

def gen_gpa(degree_type: str) -> str:
    if degree_type == "PhD":
        return f"{RNG.uniform(3.50, 4.00):.2f}"
    return f"{RNG.uniform(3.00, 3.80):.2f}"

def gen_funding(degree_type: str, study_method: str) -> int:
    if degree_type == "PhD":
        if study_method == "Full-time":
            return RNG.choices([FUND_FULL, FUND_RA, FUND_TA, FUND_SELF],
                               weights=[55, 20, 10, 15])[0]
        else:
            return RNG.choices([FUND_EMPLOY, FUND_SELF, FUND_PART, FUND_FULL],
                               weights=[35, 30, 20, 15])[0]
    else:  # Master
        if study_method == "Full-time":
            return RNG.choices([FUND_SELF, FUND_PART, FUND_FULL, FUND_TA],
                               weights=[35, 30, 25, 10])[0]
        else:
            return RNG.choices([FUND_EMPLOY, FUND_SELF, FUND_PART],
                               weights=[50, 30, 20])[0]

def gen_work(study_method: str):
    """Returns (has_external_work: bool, weekly_work_hours: float)."""
    if study_method == "Part-time":
        works = RNG.random() < 0.82
        hrs   = round(RNG.uniform(20, 40), 1) if works else 0.0
    else:
        works = RNG.random() < 0.15
        hrs   = round(RNG.uniform(5, 15), 1) if works else 0.0
    return works, hrs

def gen_research_group(degree_type: str) -> bool:
    if degree_type == "PhD":
        return RNG.random() < 0.75
    return RNG.random() < 0.30

def gen_family_support() -> int:
    return RNG.choices([2, 3, 4, 5], weights=[10, 30, 40, 20])[0]

def gen_cross_discipline(degree_type: str) -> bool:
    prob = 0.22 if degree_type == "PhD" else 0.15
    return RNG.random() < prob

def gen_email(name: str, sid_num: str) -> str:
    parts = re.sub(r"[^a-zA-Z\s]", "", name).lower().split()
    if len(parts) >= 2:
        return f"{parts[0]}.{parts[-1]}@student.mmu.edu.my"
    return f"{sid_num.lower()}@student.mmu.edu.my"


# ══════════════════════════════════════════════════════════════════════
# READ EXCEL
# ══════════════════════════════════════════════════════════════════════
wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)

ws_input = wb["Data_input"]
ws_coll  = wb["data_coll"]
h1 = [c.value for c in ws_input[1]]
h2 = [c.value for c in ws_coll[1]]
students_input = [dict(zip(h1, r)) for r in ws_input.iter_rows(min_row=2, values_only=True) if r[0]]
students_coll  = [dict(zip(h2, r)) for r in ws_coll.iter_rows(min_row=2, values_only=True) if r[0]]
wb.close()

assert len(students_input) == len(students_coll)

# Fix duplicate ID
for lst in (students_input, students_coll):
    if lst[-1]["Student_ID"] == lst[-2]["Student_ID"]:
        lst[-1]["Student_ID"] = "250PS298RL"

merged = []
for inp, col in zip(students_input, students_coll):
    assert inp["Student_ID"] == col["Student_ID"]
    merged.append({**inp, **col})


# ══════════════════════════════════════════════════════════════════════
# BUILD SQL
# ══════════════════════════════════════════════════════════════════════
lines = []
A = lines.append

A("-- ============================================")
A("-- SEED DATA — auto-generated from Excel + realistic demo values")
A(f"-- Source: PHD_data - Copy.xlsx")
A(f"-- Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}")
A("-- Excluded: marital_status, num_children, youngest_child_age")
A("-- ============================================")
A("")
A("USE datatrain;")
A("")

# ─── Regions ───
A("INSERT INTO country_region (region_name) VALUES")
regions = ["East Asia","South Asia","Southeast Asia","Middle East",
           "Europe","Africa","Americas"]
A(",\n".join(f"({q(r)})" for r in regions) + ";")
A("")

# ─── Countries ───
A("INSERT INTO country (country_name, region_id) VALUES")
countries = [
    ("China",1),("Japan",1),("South Korea",1),
    ("India",2),("Pakistan",2),("Bangladesh",2),
    ("Malaysia",3),("Indonesia",3),("Thailand",3),("Vietnam",3),
    ("Saudi Arabia",4),("Iran",4),
    ("United Kingdom",5),("Germany",5),
    ("Nigeria",6),("Kenya",6),
    ("United States",7),("Brazil",7),
]
A(",\n".join(f"({q(c)},{rid})" for c,rid in countries) + ";")
A("")

# ─── Faculties ───
A("INSERT INTO faculty (faculty_description) VALUES")
A(",\n".join(q(f) for f in [
    "Faculty of AI and Engineering","Faculty of Business",
    "Faculty of Applied Comm","Faculty of Law",
]) + ";")
A("")

# ─── Programs ───
A("INSERT INTO program (program_short_desc, program_description, faculty_id) VALUES")
A(",\n".join(f"({q(s)},{q(d_)},{fid})" for s,d_,fid in [
    ("Ph.D. (Eng)",           "Doctor of Philosophy (Engineering)",   1),
    ("Ph.D. (Mgmt.)",         "Doctor of Philosophy (Management)",    2),
    ("Master (Mgmt)",         "Master of Management",                 2),
    ("Ph.D. (Communication)", "Doctor of Philosophy (Communication)", 3),
    ("Master (Communication)","Master of Communication",              3),
    ("Master (Law)",          "Master of Law",                        4),
]) + ";")
A("")

# ─── Disciplines ───
A("INSERT INTO discipline (discipline_name, discipline_group) VALUES")
A(",\n".join(f"({q(n)},{q(g)})" for n,g in [
    ("Engineering","STEM"),("Management","Business"),
    ("Communication","Social Science"),("Law","Social Science"),
    ("Data Science","STEM"),("Finance","Business"),
]) + ";")
A("")

# ─── Funding types ───
A("INSERT INTO funding_type (funding_name) VALUES")
A(",\n".join(q(f) for f in [
    "Full Scholarship","Partial Scholarship","Self-funded",
    "Teaching Assistant","Research Assistant","Employer-sponsored",
]) + ";")
A("")

# ─── Campus ───
A("INSERT INTO campus (campus_name) VALUES")
A("('CYBER'),('MLAKA');")
A("")

# ─── Supervisors ───
A("-- Supervisors (password: 'password123'  bcrypt hash)")
A("INSERT INTO supervisor (staff_id, name, email, faculty_id, role, password_hash) VALUES")
PWD = "$2b$12$LJ3m4ys3Gz8Kvf0fRsRXB.JpGpx/A6WfFkYzS.3VqYCxLDzFqzfO6"
A(",\n".join(f"({q(s)},{q(n)},{q(e)},{f},{q(r)},{q(PWD)})" for s,n,e,f,r in [
    ("ADM001",   "Dr. Admin User",     "admin@university.edu",      1,"Both"),
    ("MU081217", "Dr. Ahmad Rahman",   "ahmad@university.edu",      1,"Supervisor"),
    ("MU092301", "Dr. Sarah Chen",     "sarah.chen@university.edu", 2,"Supervisor"),
    ("MU103456", "Prof. James Wilson", "j.wilson@university.edu",   3,"Supervisor"),
    ("MU114567", "Dr. Fatimah Idris",  "fatimah@university.edu",    4,"Supervisor"),
]) + ";")
A("")

# ══════════════════════════════════════════════════════════════════════
# STUDENTS — full column set including generated demo fields
# ══════════════════════════════════════════════════════════════════════
A("-- Students (40 from Excel + generated demo fields)")
A("INSERT INTO student (")
A("    student_id_number, student_name, email, campus_id,")
A("    campus_code, admit_term, admit_term_begin_date, expected_grad_term, program_status,")
A("    gender, date_of_birth, country_id,")
A("    program_id, degree_type, enrollment_date, study_method,")
A("    entry_gpa, is_cross_discipline,")
A("    funding_id, has_external_work, weekly_work_hours,")
A("    in_research_group, family_support")
A(") VALUES")

student_rows = []
student_ids_ordered = []
student_meta = []   # store generated fields for PPM expansion later

for rec in merged:
    sid_num  = rec["Student_ID"]
    name     = str(rec["Name"]).strip()
    campus_s = str(rec.get("Campus","CYBER")).strip().upper()
    campus_id  = CAMPUS_MAP.get(campus_s, 1)
    campus_code = rec.get("CAMPUS_ID")
    admit_term  = rec.get("Admint_Term")
    admit_begin = rec.get("AdmitTermBeginDate")
    exp_grad    = rec.get("ExpectedGradTerm")
    prog_status = rec.get("Program_Status","Active in Program")
    enroll_date = rec.get("Admission_Date")
    prog_short  = str(rec.get("Prog_short_Des","")).strip()
    prog_id, degree_type = PROGRAM_MAP.get(prog_short, (1,"PhD"))
    study_raw  = str(rec.get("Study_Method","Full-Time")).strip().lower()
    study_method = STUDY_METHOD_MAP.get(study_raw,"Full-time")

    # ── Generated demo fields ──
    gender      = infer_gender(name)
    dob         = gen_dob(enroll_date, degree_type)
    country_id  = infer_country(name)
    email       = gen_email(name, sid_num)
    gpa         = gen_gpa(degree_type)
    cross       = gen_cross_discipline(degree_type)
    funding_id  = gen_funding(degree_type, study_method)
    works, hrs  = gen_work(study_method)
    in_rg       = gen_research_group(degree_type)
    fam_sup     = gen_family_support()

    student_rows.append(
        f"({q(sid_num)},{q(name)},{q(email)},{campus_id},"
        f"{q(campus_code)},{admit_term},{d(admit_begin)},{exp_grad},{q(prog_status)},"
        f"{q(gender)},{d(dob)},{country_id},"
        f"{prog_id},{q(degree_type)},{d(enroll_date)},{q(study_method)},"
        f"{gpa},{b(cross)},"
        f"{funding_id},{b(works)},{hrs},"
        f"{b(in_rg)},{fam_sup})"
    )
    student_ids_ordered.append(sid_num)
    student_meta.append({
        "sid_num": sid_num, "name": name, "degree_type": degree_type,
        "study_method": study_method, "enroll_date": enroll_date,
    })

A(",\n".join(student_rows) + ";")
A("")
sid_to_db_id = {sid: i+1 for i, sid in enumerate(student_ids_ordered)}

# ══════════════════════════════════════════════════════════════════════
# MILESTONES
# ══════════════════════════════════════════════════════════════════════
A("-- Student Milestones (expected dates from Excel; actual = real past dates only)")
A("INSERT INTO student_milestone")
A("    (student_id, milestone_id, earliest_start_date, expected_date, actual_date, status)")
A("VALUES")

milestone_rows = []
for rec in merged:
    sid_num  = rec["Student_ID"]
    db_id    = sid_to_db_id[sid_num]
    rpd_exp  = rec.get("RPD_MAXDate")
    rpd_act  = rec.get("RPD_ActualDate")
    pub_exp  = rec.get("PublicationRequirement_MAXDate")
    pub_act  = rec.get("PublicationRequirement_paper1_ActualDay（PHD/Master）")
    ts_start = rec.get("ThesisSem_StartDate")
    ts_exp   = rec.get("ThesisSem_MAXDate")
    ts_act   = rec.get("ThesisSem_ActualDate")

    # Only mark actual as set if it's a real past date
    def past_only(dt):
        if dt is None: return None
        d2 = dt.date() if isinstance(dt, datetime) else dt
        return d2 if d2 <= TODAY else None

    milestone_rows.append(
        f"({db_id},1,NULL,{d(rpd_exp)},{d(past_only(rpd_act))},'{milestone_actual_status(rpd_exp, rpd_act)}')"
    )
    milestone_rows.append(
        f"({db_id},3,NULL,{d(pub_exp)},{d(past_only(pub_act))},'{milestone_actual_status(pub_exp, pub_act)}')"
    )
    milestone_rows.append(
        f"({db_id},4,{d(ts_start)},{d(ts_exp)},{d(past_only(ts_act))},'{milestone_actual_status(ts_exp, ts_act)}')"
    )

A(",\n".join(milestone_rows) + ";")
A("")

# ══════════════════════════════════════════════════════════════════════
# STUDENT-SUPERVISOR ASSIGNMENTS
# ══════════════════════════════════════════════════════════════════════
A("INSERT INTO student_supervisor (student_id, supervisor_id, role, assigned_date) VALUES")
prog_fac = {1:1, 2:2, 3:2, 4:3, 5:3, 6:4}
FAC_SUP  = {1:2, 2:3, 3:4, 4:5}
ss_rows  = []
for rec in merged:
    sid_num  = rec["Student_ID"]
    db_id    = sid_to_db_id[sid_num]
    prog_id, _ = PROGRAM_MAP.get(str(rec.get("Prog_short_Des","")).strip(), (1,"PhD"))
    sup_id   = FAC_SUP.get(prog_fac.get(prog_id, 1), 2)
    enroll   = rec.get("Admission_Date")
    enroll_s = enroll.strftime("%Y-%m-%d") if isinstance(enroll, datetime) else str(enroll)
    ss_rows.append(f"({db_id},{sup_id},'Main','{enroll_s}')")
A(",\n".join(ss_rows) + ";")
A("")

# ══════════════════════════════════════════════════════════════════════
# PPM RECORDS — full coverage: all 40 students × 2 cycles (2025 H1 & H2)
# ══════════════════════════════════════════════════════════════════════
A("-- PPM Records — 2 cycles per student (2025 H1 cycle 1, H2 cycle 2)")
A("-- PhD students have higher US rate for realistic ML variance")
A("INSERT INTO ppm_record")
A("    (student_id, ppm_year, ppm_cycle, result, verify_status, verified_by_id, verified_by_name)")
A("VALUES")

# Supervisor verifier map by program
prog_verifier = {
    1: ("MU081217", "Dr. Ahmad Rahman"),
    2: ("MU092301", "Dr. Sarah Chen"),
    3: ("MU092301", "Dr. Sarah Chen"),
    4: ("MU103456", "Prof. James Wilson"),
    5: ("MU103456", "Prof. James Wilson"),
    6: ("MU114567", "Dr. Fatimah Idris"),
}

ppm_rows = []
for rec in merged:
    sid_num  = rec["Student_ID"]
    db_id    = sid_to_db_id[sid_num]
    prog_id, degree_type = PROGRAM_MAP.get(str(rec.get("Prog_short_Des","")).strip(), (1,"PhD"))
    vid, vnm = prog_verifier.get(prog_id, ("MU081217", "Dr. Ahmad Rahman"))

    # US probability: PhD harder, part-time slightly harder
    study_raw   = str(rec.get("Study_Method","Full-Time")).strip().lower()
    study_method = STUDY_METHOD_MAP.get(study_raw,"Full-time")
    base_us_prob = 0.20 if degree_type == "PhD" else 0.10
    if study_method == "Part-time": base_us_prob += 0.08

    for cycle in (1, 2):
        result = "US" if RNG.random() < base_us_prob else "S"
        ppm_rows.append(
            f"({db_id},2025,{cycle},{q(result)},'Y',{q(vid)},{q(vnm)})"
        )

A(",\n".join(ppm_rows) + ";")
A("")

# ══════════════════════════════════════════════════════════════════════
# GRADUATION OUTCOMES
# ══════════════════════════════════════════════════════════════════════
A("-- Graduation Outcomes (all 40 students — real ExpectedEnd_StudyDate)")
A("INSERT INTO graduation_outcome")
A("    (student_id, expected_end_date, actual_end_date, is_delayed, final_status)")
A("VALUES")
go_rows = []
for rec in merged:
    db_id   = sid_to_db_id[rec["Student_ID"]]
    exp_end = rec.get("ExpectedEnd_StudyDate")
    go_rows.append(f"({db_id},{d(exp_end)},NULL,FALSE,'Ongoing')")
A(",\n".join(go_rows) + ";")
A("")

# ─── Examiners ───
A("INSERT INTO examiner (examiner_name, institution, is_external) VALUES")
A("('Prof. Tan Ah Kow','University of Malaya',TRUE),")
A("('Dr. Lee Wei Ming','NUS',TRUE),")
A("('Prof. Yamamoto','University of Tokyo',TRUE);")
A("")

# ══════════════════════════════════════════════════════════════════════
# WRITE OUTPUT
# ══════════════════════════════════════════════════════════════════════
sql_content = "\n".join(lines) + "\n"
with open(OUTPUT_PATH, "w", encoding="utf-8") as f:
    f.write(sql_content)

# ── Summary ──
print(f"✓ Generated {OUTPUT_PATH}")
print(f"  Students        : {len(merged)}")
print(f"  Milestone rows  : {len(milestone_rows)}")
print(f"  PPM records     : {len(ppm_rows)}  (2 cycles × {len(merged)} students)")
print(f"  Graduation rows : {len(go_rows)}")

# Quick stats on generated data
from collections import Counter
genders   = []
countries = []
works_cnt = 0
for row in student_rows:
    if "'Male'"   in row: genders.append("Male")
    elif "'Female'" in row: genders.append("Female")
    if "TRUE,0" in row or "TRUE," in row: pass  # rough check
print(f"\nGenerated field samples (first 5 students):")
for i, rec in enumerate(merged[:5]):
    sid  = rec["Student_ID"]
    name = str(rec["Name"]).strip()[:20]
    row  = student_rows[i]
    # extract gender, country from the row string (quick parse)
    g = "Male" if "'Male'" in row else "Female"
    print(f"  {name:<22} gender={g}  email={gen_email(name,sid)}")

ppm_us = sum(1 for r in ppm_rows if "'US'" in r)
print(f"\nPPM US results: {ppm_us}/{len(ppm_rows)} ({ppm_us/len(ppm_rows)*100:.0f}%)")

ms_overdue   = sum(1 for r in milestone_rows if "'Overdue'" in r)
ms_completed = sum(1 for r in milestone_rows if "'Completed'" in r)
ms_pending   = sum(1 for r in milestone_rows if "'Pending'" in r)
print(f"Milestone status: {ms_completed} Completed, {ms_pending} Pending, {ms_overdue} Overdue")
